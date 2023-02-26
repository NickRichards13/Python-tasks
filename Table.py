import string
import openpyxl
from openpyxl.styles import NamedStyle, Side, Border


def createNewTable() :
    # Ввод числа N.
    n = int(input())

    # Создание таблицы с N листами.
    book = openpyxl.Workbook()
    book.remove(book.active)
    for i in range(n):
        #Задание имён листам: sheet_N = book.create_sheet("N")
        exec("sheet_" + str(i+1) + " = book.create_sheet(\"" + str(i+1) + "\")")
    
    # Создание именного стиля границ - style_1.
    style_1 = NamedStyle(name = "style_1")
    side = Side(border_style="thin")
    style_1.border = Border(top = side, bottom = side, left = side, right = side)

    # Добавление границ и данных в нужные ячейки.
    eng_alp = string.ascii_uppercase    # Английский алфавит прописными буквами
    column = eng_alp[n-1:2*n-1] # Выбор столбцов для границы
    for sheet in book.worksheets:
        count = 1
        for r in range(n,2*n): # Выбор строк для границы
            for c in column:
                exec("sheet['" + str(c) + str(r) + "'].style = style_1")
                exec("sheet['" + str(c) + str(r) + "'] = count")
                count += 1

    # Сохранение изменений в файле.
    book.save("file.xls")

createNewTable()